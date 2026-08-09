from __future__ import annotations

import csv
import hashlib
import io
import json
import uuid
from datetime import UTC, date, datetime

import pyzipper
import pypdfium2
import pytest
from openpyxl import load_workbook

from app.models.account import Account
from app.models.app_setting import AppSetting
from app.models.audit_session import AuditSession
from app.models.transaction import Transaction
from app.seed import SAVINGS_ACCOUNT_ID

TAX_PACK_PASSPHRASE = "Correct-Horse-Archive-2026"


def _activate_pro(db):
    values = {
        "license_tier": "pro",
        "license_status": "active",
        "license_verified_at": datetime.now(UTC).isoformat(),
    }
    for key, value in values.items():
        setting = db.query(AppSetting).filter_by(key=key).first()
        if setting:
            setting.value = value
        else:
            db.add(AppSetting(key=key, value=value))
    db.commit()


def _transaction(
    db,
    *,
    txn_date,
    merchant,
    amount,
    txn_type,
    category=None,
    confidence=None,
    reconciled=False,
    is_transfer=False,
    is_income=False,
    status="settled",
    checksum=None,
    raw_text=None,
):
    transaction = Transaction(
        id=str(uuid.uuid4()),
        date=txn_date,
        raw_text=raw_text or f"Statement: {merchant}",
        merchant_raw=merchant,
        merchant_normalized=merchant,
        amount=amount,
        type=txn_type,
        instrument="statement",
        account_id=SAVINGS_ACCOUNT_ID,
        category=category,
        confidence=confidence,
        classification_source="test" if category else None,
        source="statement_upload",
        reconciled=reconciled,
        is_transfer=is_transfer,
        is_income=is_income,
        status=status,
        checksum_canonical=checksum,
    )
    db.add(transaction)
    return transaction


def test_tax_pack_requires_paid_license(auth_client):
    response = auth_client.post(
        "/api/v1/reports/fy/pack",
        json={"start_year": 2025, "passphrase": TAX_PACK_PASSPHRASE},
    )
    assert response.status_code == 403


@pytest.mark.parametrize("passphrase", ["short", "Archive-Password-2026\n"])
def test_tax_pack_rejects_unsafe_passphrase(auth_client, db_session, passphrase):
    _activate_pro(db_session)
    response = auth_client.post(
        "/api/v1/reports/fy/pack",
        json={"start_year": 2025, "passphrase": passphrase},
    )
    assert response.status_code == 422


def test_tax_pack_contents_hashes_workbook_pdf_and_warnings(
    auth_client,
    db_session,
):
    _activate_pro(db_session)
    account = db_session.get(Account, SAVINGS_ACCOUNT_ID)
    assert account.nickname.startswith("Example ")
    _transaction(
        db_session,
        txn_date=date(2025, 4, 1),
        merchant="SALARY",
        amount=100000,
        txn_type="credit",
        category="INCOME",
        confidence=1,
        reconciled=True,
        is_income=True,
    )
    _transaction(
        db_session,
        txn_date=date(2025, 4, 2),
        merchant="=DANGEROUS FORMULA",
        amount=750,
        txn_type="debit",
        confidence=0.4,
        checksum="duplicate-risk",
    )
    _transaction(
        db_session,
        txn_date=date(2025, 4, 2),
        merchant="DUPLICATE COPY",
        amount=750,
        txn_type="debit",
        category="MISCELLANEOUS",
        confidence=0.8,
        checksum="duplicate-risk",
    )
    _transaction(
        db_session,
        txn_date=date(2025, 4, 3),
        merchant="OWN ACCOUNT TRANSFER",
        amount=5000,
        txn_type="debit",
        category="TRANSFERS",
        confidence=1,
        reconciled=True,
        is_transfer=True,
    )
    _transaction(
        db_session,
        txn_date=date(2025, 4, 4),
        merchant="CARD REVERSAL",
        amount=500,
        txn_type="credit",
        status="reversed",
        reconciled=True,
    )
    _transaction(
        db_session,
        txn_date=date(2025, 4, 5),
        merchant="alice@example.com alice@okhdfcbank 9876543210 123456789012",
        amount=12500,
        txn_type="credit",
        raw_text=(
            "UPI from alice@example.com via alice@okhdfcbank phone 9876543210 "
            "reference 123456789012"
        ),
    )
    db_session.add(
        AuditSession(
            period_year=2025,
            period_month=4,
            status="finalized",
        )
    )
    db_session.commit()

    response = auth_client.post(
        "/api/v1/reports/fy/pack",
        json={"start_year": 2025, "passphrase": TAX_PACK_PASSPHRASE},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/zip"
    assert response.headers["cache-control"] == "no-store"
    assert response.headers["x-content-type-options"] == "nosniff"
    assert "godfin_ca_tax_pack_fy2025-26.zip" in response.headers[
        "content-disposition"
    ]
    assert TAX_PACK_PASSPHRASE.encode() not in response.content

    with pyzipper.AESZipFile(io.BytesIO(response.content)) as locked_archive:
        with pytest.raises(RuntimeError):
            locked_archive.read("manifest.json")

    with pyzipper.AESZipFile(io.BytesIO(response.content)) as wrong_archive:
        wrong_archive.setpassword(b"This-Is-The-Wrong-Password")
        with pytest.raises(RuntimeError):
            wrong_archive.read("manifest.json")

    with pyzipper.AESZipFile(io.BytesIO(response.content)) as archive:
        archive.setpassword(TAX_PACK_PASSPHRASE.encode())
        assert archive.infolist()
        assert all(info.flag_bits & 0x1 for info in archive.infolist())
        names = set(archive.namelist())
        workbook_name = "godfin_ca_fy2025-26.xlsx"
        csv_name = "godfin_ca_fy2025-26_transactions.csv"
        guide_name = "filing_guide_ay2026_27_v1.1.pdf"
        required = {
            workbook_name,
            csv_name,
            "manifest.json",
            "reconciliation_summary.json",
            guide_name,
        }
        assert required <= names

        manifest = json.loads(archive.read("manifest.json"))
        reconciliation = json.loads(
            archive.read("reconciliation_summary.json")
        )
        assert manifest["financial_year"] == "FY 2025-26"
        assert manifest["assessment_year"] == "AY 2026-27"
        assert manifest["schema_version"] == "2.0"
        assert manifest["encryption"]["algorithm"] == "AES-256"
        assert manifest["encryption"]["passphrase_stored"] is False
        assert manifest["privacy"]["raw_narration_included"] is False
        assert reconciliation["readiness"] == "review_required"
        assert reconciliation["unclassified_count"] == 3
        assert reconciliation["duplicate_risk_count"] == 2
        assert reconciliation["period_warning_count"] > 0
        for filename, evidence in manifest["files"].items():
            content = archive.read(filename)
            assert hashlib.sha256(content).hexdigest() == evidence["sha256"]
            assert len(content) == evidence["bytes"]

        workbook = load_workbook(
            io.BytesIO(archive.read(workbook_name)),
            data_only=False,
        )
        assert {
            "Metadata",
            "Reconciliation Summary",
            "Masked Accounts",
            "Transactions",
            "Income Review",
            "Expense Tax Review",
            "Transfers Other Credits",
            "Data Quality Exceptions",
            "Period Completeness",
            "Missing Evidence",
            "Filing Steps",
        } <= set(workbook.sheetnames)
        assert "Economic meaning" in [
            cell.value for cell in workbook["Transactions"][1]
        ]
        assert "Narration (locally redacted)" in [
            cell.value for cell in workbook["Transactions"][1]
        ]
        assert "Raw narration" not in [
            cell.value for cell in workbook["Transactions"][1]
        ]
        transactions = workbook["Transactions"]
        assert transactions.freeze_panes == "A2"
        assert transactions.auto_filter.ref
        headers = {
            cell.value: index
            for index, cell in enumerate(transactions[1], start=1)
        }
        amount_cell = transactions.cell(row=2, column=headers["Amount (INR)"])
        assert isinstance(amount_cell.value, (int, float))
        assert "₹" in amount_cell.number_format
        merchant_values = [
            transactions.cell(row=row, column=headers["Merchant"]).value
            for row in range(2, transactions.max_row + 1)
        ]
        assert "'=DANGEROUS FORMULA" in merchant_values
        assert not any("alice@example.com" in str(value) for value in merchant_values)
        assert not any("alice@okhdfcbank" in str(value) for value in merchant_values)
        assert not any("9876543210" in str(value) for value in merchant_values)
        assert not any("123456789012" in str(value) for value in merchant_values)
        assert any("[EMAIL REDACTED]" in str(value) for value in merchant_values)
        income_review = workbook["Income Review"]
        income_headers = {
            cell.value: index
            for index, cell in enumerate(income_review[1], start=1)
        }
        income_merchants = [
            income_review.cell(row=row, column=income_headers["Merchant"]).value
            for row in range(2, income_review.max_row + 1)
        ]
        assert income_merchants == ["SALARY"]
        period_sheet = workbook["Period Completeness"]
        period_headers = [cell.value for cell in period_sheet[1]]
        assert "Authoritative audit session" in period_headers
        assert "Review required" in period_headers
        assert period_sheet.max_row == 13
        summary = workbook["Reconciliation Summary"]
        assert summary["B2"].data_type == "f"
        workbook.close()

        rows = list(
            csv.DictReader(
                io.StringIO(
                    archive.read(csv_name)
                    .decode("utf-8-sig")
                )
            )
        )
        assert any(
            row["Merchant"].startswith("'=DANGEROUS")
            for row in rows
        )
        serialized_rows = json.dumps(rows)
        assert "alice@example.com" not in serialized_rows
        assert "alice@okhdfcbank" not in serialized_rows
        assert "9876543210" not in serialized_rows
        assert "123456789012" not in serialized_rows
        assert "[EMAIL REDACTED]" in serialized_rows

        guide = archive.read(guide_name)
        assert guide.startswith(b"%PDF")
        document = pypdfium2.PdfDocument(guide)
        assert len(document) >= 1
        rendered = document[0].render(scale=0.5).to_pil()
        assert rendered.width > 0 and rendered.height > 0
        document.close()


def test_existing_financial_year_csv_and_json_are_preserved(
    auth_client,
    db_session,
):
    _activate_pro(db_session)
    csv_response = auth_client.get(
        "/api/v1/reports/fy?start_year=2025&format=csv"
    )
    json_response = auth_client.get(
        "/api/v1/reports/fy?start_year=2025&format=json"
    )
    assert csv_response.status_code == 200
    assert json_response.status_code == 200
    assert "transactions" in json_response.json()
