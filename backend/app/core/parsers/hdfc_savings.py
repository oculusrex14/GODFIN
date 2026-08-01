from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from typing import Optional

import pdfplumber
from openpyxl import load_workbook

from app.core.parsers.base import StatementParserPlugin
from app.core.statement_parser import (
    StatementParseResult,
    _append_strict_savings_txn,
    _parse_hdfc_savings_statement,
    _parse_statement_date,
    _validate_savings_controls,
    parse_statement_xls,
)

logger = logging.getLogger(__name__)


def _detect(text: str) -> bool:
    normalized = text.lower()
    return "hdfc" in normalized and (
        "savings account" in normalized or "statement of account" in normalized
    )


def _parse_pdf(
    contents: bytes,
    _file_format: str,
    password: Optional[str],
) -> StatementParseResult:
    result = StatementParseResult(
        statement_type="hdfc_savings",
        parser_profile="hdfc_savings",
        recognized=True,
    )
    try:
        pdf = pdfplumber.open(io.BytesIO(contents), password=password)
    except Exception as exc:
        result.errors.append(f"Failed to open PDF: {exc}")
        return result

    try:
        _parse_hdfc_savings_statement(pdf, result)
    except Exception as exc:
        result.errors.append(f"Parse error: {exc}")
    finally:
        pdf.close()
    return result


def _date_value(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return _parse_statement_date(str(value or "").strip())


def _amount_value(value: object) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(cleaned) if cleaned else None
    except ValueError:
        return None


def _parse_xlsx(contents: bytes) -> StatementParseResult:
    result = StatementParseResult(
        statement_type="hdfc_savings",
        parser_profile="hdfc_savings",
    )
    try:
        workbook = load_workbook(
            io.BytesIO(contents),
            read_only=True,
            data_only=True,
        )
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as exc:
        result.errors.append(f"Failed to open XLSX: {exc}")
        return result

    header_index: Optional[int] = None
    columns: dict[str, int] = {}
    try:
        metadata_text = " ".join(
            str(value or "")
            for row in rows[:30]
            for value in row
        ).upper()
        if "HDFC" not in metadata_text:
            result.errors.append("HDFC bank fingerprint was not found in XLSX metadata")
            return result

        for index, row in enumerate(rows[:30]):
            values = [str(value or "").strip() for value in row]
            joined = " ".join(values)
            upper = joined.upper()

            if "STATEMENT FROM" in upper or "FROM :" in upper:
                dates = re.findall(r"\d{2}/\d{2}/\d{2,4}", joined)
                if len(dates) >= 2:
                    result.period_start = _parse_statement_date(dates[0])
                    result.period_end = _parse_statement_date(dates[1])

            if "NARRATION" not in upper or "DATE" not in upper:
                continue

            header_index = index
            for column, value in enumerate(values):
                cell = value.upper()
                if cell == "DATE":
                    columns["date"] = column
                elif "NARRATION" in cell:
                    columns["narration"] = column
                elif "CHQ" in cell or "REF" in cell:
                    columns["ref"] = column
                elif "VALUE" in cell:
                    columns["value_date"] = column
                elif "WITHDRAWAL" in cell:
                    columns["withdrawal"] = column
                elif "DEPOSIT" in cell:
                    columns["deposit"] = column
                elif "CLOSING" in cell or "BALANCE" in cell:
                    columns["balance"] = column
            break

        if header_index is None:
            result.errors.append("Could not find header row in XLSX")
            return result

        required_columns = {"date", "narration", "withdrawal", "deposit", "balance"}
        missing_columns = sorted(required_columns - set(columns))
        if missing_columns:
            result.errors.append(
                "Missing required XLSX columns: " + ", ".join(missing_columns),
            )
            return result
        result.recognized = True

        for row in rows[header_index + 1 :]:
            first = str(row[0] or "").strip() if row else ""
            if first.startswith("*"):
                continue

            date_index = columns["date"]
            if date_index >= len(row):
                continue
            transaction_date = _date_value(row[date_index])
            if transaction_date is None:
                continue

            narration_index = columns["narration"]
            narration = (
                str(row[narration_index] or "").strip()
                if narration_index < len(row)
                else ""
            )
            if not narration:
                continue

            def cell(name: str) -> object:
                position = columns[name]
                return row[position] if position < len(row) else None

            _append_strict_savings_txn(
                {
                    "date": transaction_date,
                    "narration": narration,
                    "ref": str(cell("ref") or "").strip(),
                    "value_date": cell("value_date"),
                    "withdrawal": _amount_value(cell("withdrawal")),
                    "deposit": _amount_value(cell("deposit")),
                    "balance": _amount_value(cell("balance")),
                },
                result.transactions,
                result.errors,
                row_label=f"XLSX row {header_index + 2 + len(result.transactions)}",
            )
    finally:
        workbook.close()

    if not result.transactions and not result.errors:
        result.errors.append("No transactions found in XLSX")
    _validate_savings_controls(result)
    logger.info("XLSX parser: found %s transactions", len(result.transactions))
    return result


def _parse(
    contents: bytes,
    file_format: str,
    password: Optional[str],
) -> StatementParseResult:
    if file_format == "pdf":
        return _parse_pdf(contents, file_format, password)
    if file_format == "xls":
        return parse_statement_xls(contents)
    if file_format == "xlsx":
        return _parse_xlsx(contents)
    return StatementParseResult(
        statement_type="hdfc_savings",
        errors=[f"Unsupported HDFC savings format: {file_format}"],
    )


PARSER = StatementParserPlugin(
    profile="hdfc_savings",
    bank="HDFC",
    account_type="savings",
    statement_type="hdfc_savings",
    formats=("pdf", "xls", "xlsx"),
    detect_text=_detect,
    parse=_parse,
)
