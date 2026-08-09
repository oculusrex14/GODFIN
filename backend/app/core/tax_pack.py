from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from collections import Counter
from datetime import UTC, date, datetime, timedelta
from typing import Any, Iterable

import pyzipper
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.account import Account
from app.models.audit_session import AuditSession
from app.models.transaction import Transaction
from app.core.money import money_decimal
from app.core.transaction_semantics import (
    TransactionSemantic,
    is_spending,
    is_verified_income,
    semantic_type_for,
)

TAX_PACK_SCHEMA_VERSION = "2.0"
TAX_GUIDE_VERSION = "1.1"
LOW_CONFIDENCE_THRESHOLD = 0.75
MIN_TAX_PACK_PASSPHRASE_LENGTH = 12
MAX_TAX_PACK_PASSPHRASE_LENGTH = 128
OFFICIAL_DOWNLOADS_URL = (
    "https://www.incometax.gov.in/iec/foportal/downloads/income-tax-returns"
)
OFFICIAL_AIS_URL = (
    "https://www.incometax.gov.in/iec/foportal/help/all-topics/"
    "e-filing-services/ais%20-%20annual%20information%20statement-faqs"
)
INR_FORMAT = '[$₹-en-IN]#,##0.00'
DATE_FORMAT = "dd-mm-yyyy"
HEADER_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
TRANSACTION_COLUMNS = [
    "Transaction ID",
    "Date",
    "Time",
    "Account",
    "Merchant",
    "Narration (locally redacted)",
    "Amount (INR)",
    "Type",
    "Instrument",
    "Category",
    "Subcategory",
    "Classification confidence",
    "Classification source",
    "Status",
    "Transfer",
    "Recurring",
    "Income",
    "Economic meaning",
    "Source",
    "Reconciled",
    "Duplicate risk",
    "Tags",
    "Notes (locally redacted)",
]

_EMAIL_PATTERN = re.compile(
    r"(?<![A-Z0-9._%+-])[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}(?![A-Z0-9])",
    re.IGNORECASE,
)
_UPI_PATTERN = re.compile(
    r"(?<![A-Z0-9._+-])[A-Z0-9._+-]{2,}@[A-Z][A-Z0-9._-]{1,}(?![A-Z0-9._-])",
    re.IGNORECASE,
)
_PHONE_PATTERN = re.compile(r"(?<!\d)(?:\+?91[\s-]?)?[6-9]\d{9}(?!\d)")
_LONG_NUMBER_PATTERN = re.compile(r"(?<!\d)\d{6,}(?!\d)")
_CONTROL_PATTERN = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_MAX_REDACTED_TEXT_LENGTH = 500


def _safe_text(value: Any) -> Any:
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    cleaned = value.replace("\x00", "").strip()
    if cleaned.startswith(("=", "+", "-", "@")):
        return f"'{cleaned}"
    return cleaned


def _redact_sensitive_text(value: Any) -> str:
    """Minimize common personal identifiers before a tax pack leaves GODFIN."""
    if value is None:
        return ""
    cleaned = _CONTROL_PATTERN.sub(" ", str(value)).strip()
    cleaned = _EMAIL_PATTERN.sub("[EMAIL REDACTED]", cleaned)
    cleaned = _UPI_PATTERN.sub("[UPI ID REDACTED]", cleaned)
    cleaned = _PHONE_PATTERN.sub("[PHONE REDACTED]", cleaned)
    cleaned = _LONG_NUMBER_PATTERN.sub("[REFERENCE NUMBER REDACTED]", cleaned)
    cleaned = " ".join(cleaned.split())
    if len(cleaned) > _MAX_REDACTED_TEXT_LENGTH:
        cleaned = f"{cleaned[:_MAX_REDACTED_TEXT_LENGTH].rstrip()}…"
    return str(_safe_text(cleaned))


def _masked_account(account: Account | None) -> str:
    if not account:
        return "Unknown account"
    return (
        f"{account.bank} "
        f"{account.account_type.replace('_', ' ')} ****{account.last_4_digits}"
    )


def _is_reversal(transaction: Transaction) -> bool:
    return semantic_type_for(transaction) == TransactionSemantic.REVERSAL.value


def _transaction_row(
    transaction: Transaction,
    accounts: dict[str, Account],
    duplicate_checksums: set[str],
) -> dict[str, Any]:
    return {
        "Transaction ID": transaction.id,
        "Date": transaction.date,
        "Time": transaction.time,
        "Account": _masked_account(accounts.get(transaction.account_id)),
        "Merchant": _redact_sensitive_text(
            transaction.merchant_normalized or transaction.merchant_raw
        ),
        "Narration (locally redacted)": _redact_sensitive_text(
            transaction.raw_text
        ),
        "Amount (INR)": round(float(transaction.amount), 2),
        "Type": transaction.type,
        "Instrument": transaction.instrument,
        "Category": transaction.category,
        "Subcategory": transaction.subcategory,
        "Classification confidence": (
            round(float(transaction.confidence), 4)
            if transaction.confidence is not None
            else None
        ),
        "Classification source": transaction.classification_source,
        "Status": transaction.status,
        "Transfer": bool(transaction.is_transfer),
        "Recurring": bool(transaction.is_recurring),
        "Income": bool(transaction.is_income),
        "Economic meaning": semantic_type_for(transaction),
        "Source": transaction.source,
        "Reconciled": bool(transaction.reconciled),
        "Duplicate risk": bool(
            transaction.checksum_canonical
            and transaction.checksum_canonical in duplicate_checksums
        ),
        "Tags": _redact_sensitive_text(transaction.tags),
        "Notes (locally redacted)": _redact_sensitive_text(transaction.notes),
    }


def _quality_exceptions(
    transactions: Iterable[Transaction],
    accounts: dict[str, Account],
    duplicate_checksums: set[str],
) -> list[dict[str, Any]]:
    exceptions: list[dict[str, Any]] = []
    for transaction in transactions:
        issues = []
        if not transaction.category:
            issues.append(("unclassified", "Assign and confirm a category."))
        if transaction.type == "credit" and not is_verified_income(transaction):
            issues.append(
                (
                    "unverified_credit",
                    "Confirm whether this is income, a transfer, refund, cashback, reimbursement or reversal.",
                )
            )
        if (
            transaction.confidence is not None
            and transaction.confidence < LOW_CONFIDENCE_THRESHOLD
        ):
            issues.append(
                (
                    "low_classification_confidence",
                    "Review the category and supporting narration.",
                )
            )
        if (
            transaction.checksum_canonical
            and transaction.checksum_canonical in duplicate_checksums
        ):
            issues.append(
                (
                    "duplicate_risk",
                    "Confirm that this is not duplicated across email and statement sources.",
                )
            )
        if not transaction.reconciled:
            issues.append(
                (
                    "not_reconciled",
                    "Reconcile this transaction against a bank or card statement.",
                )
            )
        if not (transaction.raw_text or transaction.notes):
            issues.append(
                (
                    "missing_narration",
                    "Attach or retain source evidence outside GODFIN.",
                )
            )
        for code, action in issues:
            exceptions.append(
                {
                    "Transaction ID": transaction.id,
                    "Date": transaction.date,
                    "Account": _masked_account(accounts.get(transaction.account_id)),
                    "Merchant": _redact_sensitive_text(
                        transaction.merchant_normalized
                        or transaction.merchant_raw
                    ),
                    "Amount (INR)": round(float(transaction.amount), 2),
                    "Issue": code,
                    "Required review": action,
                }
            )
    return exceptions


def _month_keys(start_year: int) -> list[tuple[int, int]]:
    return [
        (
            start_year if offset < 9 else start_year + 1,
            offset + 4 if offset < 9 else offset - 8,
        )
        for offset in range(12)
    ]


def _period_review_rows(
    db: Session,
    transactions: list[Transaction],
    start_year: int,
) -> list[dict[str, Any]]:
    transaction_counts = Counter(
        (transaction.date.year, transaction.date.month)
        for transaction in transactions
    )
    sessions = (
        db.query(AuditSession)
        .filter(
            AuditSession.period_year.in_([start_year, start_year + 1]),
            AuditSession.status != "discarded",
        )
        .order_by(AuditSession.created_at.desc())
        .all()
    )
    by_period: dict[tuple[int, int], list[AuditSession]] = {}
    for session in sessions:
        by_period.setdefault(
            (session.period_year, session.period_month), []
        ).append(session)

    rows = []
    for year, month in _month_keys(start_year):
        period_sessions = by_period.get((year, month), [])
        session = period_sessions[0] if len(period_sessions) == 1 else None
        issues = []
        actions = []
        if not transaction_counts[(year, month)]:
            issues.append("No transactions present")
            actions.append(
                "Confirm that the period is genuinely inactive or import the missing statement/email data."
            )
        if len(period_sessions) > 1:
            issues.append("Multiple active audit sessions")
            actions.append(
                "Resolve the audit-session conflict before relying on this period."
            )
        if not session or session.status not in {"finalized", "locked"}:
            issues.append("Month not finalized")
            actions.append(
                "Review and finalize the month before treating the pack as complete."
            )
        rows.append(
            {
                "Period": f"{month:02d}-{year}",
                "Transaction count": transaction_counts[(year, month)],
                "Authoritative audit session": session.id if session else "",
                "Audit status": session.status if session else "no audit session",
                "Finalized UTC": (
                    session.finalized_at.isoformat()
                    if session and session.finalized_at
                    else ""
                ),
                "Review required": bool(issues),
                "Issue": "; ".join(issues) if issues else "No automated period issue",
                "Required review": (
                    " ".join(actions)
                    if actions
                    else "Reconcile external tax records before filing."
                ),
            }
        )
    return rows


def _append_table(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    *,
    columns: list[str] | None = None,
) -> None:
    sheet = workbook.create_sheet(title[:31])
    columns = columns or (list(rows[0]) if rows else ["Information"])
    sheet.append(columns)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True)
    for row in rows:
        sheet.append([_safe_text(row.get(column)) for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell, column in zip(row, columns):
            if column == "Date" and isinstance(cell.value, date):
                cell.number_format = DATE_FORMAT
            elif "(INR)" in column and isinstance(cell.value, (int, float)):
                cell.number_format = INR_FORMAT
            elif column == "Classification confidence" and isinstance(
                cell.value, (int, float)
            ):
                cell.number_format = "0.0%"
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, column in enumerate(columns, start=1):
        sample = [str(column)] + [
            str(sheet.cell(row=row, column=index).value or "")
            for row in range(2, min(sheet.max_row, 80) + 1)
        ]
        width = min(52, max(12, max(len(value) for value in sample) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width


def _build_workbook(
    *,
    metadata: list[dict[str, Any]],
    account_rows: list[dict[str, Any]],
    transaction_rows: list[dict[str, Any]],
    income_rows: list[dict[str, Any]],
    expense_rows: list[dict[str, Any]],
    transfer_rows: list[dict[str, Any]],
    exception_rows: list[dict[str, Any]],
    period_rows: list[dict[str, Any]],
    evidence_rows: list[dict[str, Any]],
    filing_rows: list[dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _append_table(workbook, "Metadata", metadata, columns=["Field", "Value"])
    _append_table(
        workbook,
        "Masked Accounts",
        account_rows,
        columns=["Account", "Bank", "Type", "Active"],
    )
    _append_table(
        workbook,
        "Transactions",
        transaction_rows,
        columns=TRANSACTION_COLUMNS,
    )
    _append_table(
        workbook,
        "Income Review",
        income_rows,
        columns=TRANSACTION_COLUMNS,
    )
    _append_table(
        workbook,
        "Expense Tax Review",
        expense_rows,
        columns=[*TRANSACTION_COLUMNS, "Tax review status"],
    )
    _append_table(
        workbook,
        "Transfers Other Credits",
        transfer_rows,
        columns=TRANSACTION_COLUMNS,
    )
    _append_table(workbook, "Data Quality Exceptions", exception_rows)
    _append_table(workbook, "Period Completeness", period_rows)
    _append_table(
        workbook,
        "Missing Evidence",
        evidence_rows,
        columns=["Evidence", "Why it is needed", "Status"],
    )
    _append_table(
        workbook,
        "Filing Steps",
        filing_rows,
        columns=["Step", "Action", "Owner"],
    )

    summary = workbook.create_sheet("Reconciliation Summary", 1)
    summary.append(["Metric", "Workbook formula"])
    for cell in summary[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    summary_rows = [
        ("All transactions", "=MAX(COUNTA(Transactions!A:A)-1,0)"),
        (
            "Total income candidate amount",
            "=SUM('Income Review'!G:G)",
        ),
        (
            "Total expense review amount",
            "=SUM('Expense Tax Review'!G:G)",
        ),
        (
            "Data-quality exceptions",
            "=MAX(COUNTA('Data Quality Exceptions'!A:A)-1,0)",
        ),
        (
            "Period completeness warnings",
            "=MAX(COUNTA('Period Completeness'!A:A)-1,0)",
        ),
    ]
    for label, formula in summary_rows:
        summary.append([label, formula])
    summary.freeze_panes = "A2"
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 28
    for cell in summary["B"][1:]:
        cell.number_format = INR_FORMAT if "amount" in summary.cell(
            row=cell.row, column=1
        ).value.lower() else "0"

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _csv_bytes(rows: list[dict[str, Any]]) -> bytes:
    output = io.StringIO(newline="")
    columns = list(rows[0]) if rows else [
        "Transaction ID",
        "Date",
        "Account",
        "Merchant",
        "Amount (INR)",
        "Type",
    ]
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                column: (
                    value.strftime("%d-%m-%Y")
                    if isinstance(value, date)
                    else _safe_text(value)
                )
                for column, value in row.items()
            }
        )
    return output.getvalue().encode("utf-8-sig")


def _filing_guide_pdf(
    *,
    fy_label: str,
    ay_label: str,
    generated_at: str,
    exception_count: int,
    period_warning_count: int,
) -> bytes:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_title(f"GODFIN Indian income-tax filing guide {ay_label}")
    pdf.set_author("GODFIN")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 10, f"GODFIN Tax-Pack Guide - {ay_label}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(
        0,
        5,
        (
            f"Guide version {TAX_GUIDE_VERSION} | {fy_label} | Generated {generated_at}\n"
            "This pack organizes local transaction data. It is not an income-tax "
            "return, does not determine the correct ITR, and does not replace a "
            "Chartered Accountant. Transaction data alone cannot establish taxable "
            "income, deductions, cost basis, ownership, residency, or filing status."
        ),
        align="L",
        new_x="LMARGIN",
        new_y="NEXT",
    )

    sections = [
        (
            "Protect and share this pack safely",
            [
                "Every file in this ZIP uses AES-256 encryption. Use an AES-capable archive app if the operating system's built-in extractor cannot open it.",
                "Send the ZIP and its passphrase through different channels. Do not place the password in the same email or message as the archive.",
                "GODFIN does not store the passphrase and cannot recover it. ZIP filenames remain visible even though file contents are encrypted.",
                "Common emails, UPI IDs, phone numbers and long reference numbers are locally redacted, but dates, amounts and tax-review details remain sensitive.",
            ],
        ),
        (
            "1. Review GODFIN's completeness warnings",
            [
                f"Resolve or explain {exception_count} transaction-level exceptions.",
                f"Review {period_warning_count} period-completeness warnings.",
                "Do not treat this pack as filing-ready while unclassified, low-confidence, duplicate-risk, unreconciled, missing, or unfinalized items remain.",
            ],
        ),
        (
            "2. Collect records GODFIN cannot infer",
            [
                "Form 16 and Form 16A/16B/16C, salary and pension records.",
                "Bank interest certificates, broker capital-gain statements, contract notes, mutual-fund statements, dividend records, and foreign-asset/income records where applicable.",
                "Home-loan, rent/HRA, insurance, donation, medical, education-loan, tax-payment and deduction evidence.",
                "Business books, invoices, GST records, depreciation schedules, and profession receipts where applicable.",
            ],
        ),
        (
            "3. Download and reconcile government information",
            [
                "Download AIS and TIS from the Income Tax e-Filing portal.",
                "Check Form 26AS for TDS/TCS and tax-payment information.",
                "Compare AIS/TIS/26AS, Form 16 and supporting statements against this pack. AIS may not contain every transaction; the taxpayer remains responsible for complete and accurate reporting.",
                "Submit AIS feedback where source information is wrong and retain its acknowledgement.",
            ],
        ),
        (
            "4. Choose the correct return",
            [
                "Use the current official AY instructions. ITR-1, ITR-2, ITR-3 and ITR-4 have different eligibility rules.",
                "Business/profession income, capital gains, foreign assets/income, director or unlisted-share holdings, residency, total income and other facts can change the correct form.",
                "Ask a CA when eligibility is uncertain. GODFIN cannot choose the return from bank transactions alone.",
            ],
        ),
        (
            "5. Prepare, validate and submit",
            [
                "Use the latest portal or official Common Offline Utility for the assessment year.",
                "Enter and reconcile each income head, deduction, tax paid, loss, asset and disclosure from authoritative evidence.",
                "Run portal/utility validation, review the tax computation and bank/refund details, then submit.",
                "E-verify within the permitted period using an available method. Save the submitted return, computation, ITR-V/acknowledgement, verification confirmation and source records.",
            ],
        ),
        (
            "Official references",
            [
                f"AY utilities, forms, schemas and validations: {OFFICIAL_DOWNLOADS_URL}",
                f"AIS/TIS guidance and feedback: {OFFICIAL_AIS_URL}",
                "Verify the current assessment-year instructions on the Income Tax Department portal before filing.",
            ],
        ),
    ]
    for heading, bullets in sections:
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 12)
        pdf.multi_cell(
            0,
            6,
            heading,
            align="L",
            new_x="LMARGIN",
            new_y="NEXT",
        )
        pdf.set_font("Helvetica", "", 9)
        for item in bullets:
            pdf.multi_cell(
                0,
                5,
                f"- {item}",
                align="L",
                new_x="LMARGIN",
                new_y="NEXT",
            )
    return bytes(pdf.output())


def build_financial_year_tax_pack(
    db: Session,
    start_year: int,
    *,
    passphrase: str,
) -> bytes:
    if not (
        MIN_TAX_PACK_PASSPHRASE_LENGTH
        <= len(passphrase)
        <= MAX_TAX_PACK_PASSPHRASE_LENGTH
    ):
        raise ValueError(
            "Tax-pack passphrase must be between 12 and 128 characters"
        )
    if any(ord(character) < 32 or ord(character) == 127 for character in passphrase):
        raise ValueError("Tax-pack passphrase cannot contain control characters")

    start = date(start_year, 4, 1)
    end = date(start_year + 1, 4, 1)
    generated = datetime.now(UTC).replace(microsecond=0)
    transactions = (
        db.query(Transaction)
        .filter(
            Transaction.date >= start,
            Transaction.date < end,
            Transaction.status != "deleted",
        )
        .order_by(Transaction.date, Transaction.time, Transaction.id)
        .all()
    )
    accounts = {account.id: account for account in db.query(Account).all()}
    checksum_counts = Counter(
        transaction.checksum_canonical
        for transaction in transactions
        if transaction.checksum_canonical
    )
    duplicate_checksums = {
        checksum for checksum, count in checksum_counts.items() if count > 1
    }
    transaction_rows = [
        _transaction_row(transaction, accounts, duplicate_checksums)
        for transaction in transactions
    ]
    income_rows = [
        row
        for transaction, row in zip(transactions, transaction_rows)
        if is_verified_income(transaction)
    ]
    expense_rows = [
        {
            **row,
            "Tax review status": "Review with CA; no deductibility asserted",
        }
        for transaction, row in zip(transactions, transaction_rows)
        if is_spending(transaction)
    ]
    transfer_rows = [
        row
        for transaction, row in zip(transactions, transaction_rows)
        if (
            transaction.is_transfer
            or _is_reversal(transaction)
            or (
                transaction.type == "credit"
                and not is_verified_income(transaction)
            )
        )
    ]
    exceptions = _quality_exceptions(
        transactions, accounts, duplicate_checksums
    )
    period_rows = _period_review_rows(db, transactions, start_year)
    period_warning_count = sum(
        1 for row in period_rows if row["Review required"]
    )
    fy_label = f"FY {start_year}-{str(start_year + 1)[-2:]}"
    ay_label = f"AY {start_year + 1}-{str(start_year + 2)[-2:]}"
    total_income = money_decimal(
        sum(
            (money_decimal(row["Amount (INR)"]) for row in income_rows),
            money_decimal(0),
        )
    )
    total_expense = money_decimal(
        sum(
            (money_decimal(row["Amount (INR)"]) for row in expense_rows),
            money_decimal(0),
        )
    )

    metadata = [
        {"Field": "Schema version", "Value": TAX_PACK_SCHEMA_VERSION},
        {"Field": "Guide version", "Value": TAX_GUIDE_VERSION},
        {"Field": "Financial year", "Value": fy_label},
        {"Field": "Assessment year", "Value": ay_label},
        {"Field": "Period start", "Value": start.strftime("%d-%m-%Y")},
        {
            "Field": "Period end",
            "Value": (end - timedelta(days=1)).strftime("%d-%m-%Y"),
        },
        {"Field": "Generated UTC", "Value": generated.isoformat()},
        {
            "Field": "Readiness",
            "Value": (
                "REVIEW REQUIRED"
                if exceptions or period_warning_count
                else "No automated exceptions; external evidence still required"
            ),
        },
        {
            "Field": "Scope warning",
            "Value": (
                "Local transaction organizer only; not an ITR, tax computation, "
                "or substitute for a Chartered Accountant."
            ),
        },
        {
            "Field": "Privacy",
            "Value": (
                "AES-256 encrypted. Common identifiers in narration, merchant, tags "
                "and notes are locally redacted; ZIP filenames remain visible."
            ),
        },
        {
            "Field": "Password sharing",
            "Value": (
                "Send the passphrase to the intended CA through a different channel "
                "from the archive. GODFIN never stores it."
            ),
        },
    ]
    account_rows = [
        {
            "Account": _masked_account(account),
            "Bank": account.bank,
            "Type": account.account_type.replace("_", " "),
            "Active": bool(account.is_active),
        }
        for account in sorted(accounts.values(), key=lambda item: item.created_at)
    ]
    evidence_rows = [
        {
            "Evidence": name,
            "Why it is needed": reason,
            "Status": "Not tracked by GODFIN - verify externally",
        }
        for name, reason in [
            ("AIS and TIS", "Reconcile information reported to the department."),
            ("Form 26AS", "Reconcile TDS/TCS and tax payments."),
            ("Form 16 / salary evidence", "Confirm salary, exemptions and TDS."),
            ("Interest certificates", "Confirm savings, FD and other interest."),
            ("Capital-gain statements", "Confirm proceeds, cost basis and gains."),
            ("Deduction evidence", "Support any deduction actually claimed."),
            ("Foreign asset/income evidence", "Support applicable disclosures."),
            ("ITR acknowledgement", "Retain proof after submission and e-verification."),
        ]
    ]
    filing_rows = [
        {"Step": 1, "Action": "Resolve GODFIN exceptions and incomplete periods.", "Owner": "User"},
        {"Step": 2, "Action": "Collect external income, tax, asset and deduction evidence.", "Owner": "User / CA"},
        {"Step": 3, "Action": "Download and reconcile AIS, TIS and Form 26AS.", "Owner": "User / CA"},
        {"Step": 4, "Action": "Choose the correct ITR using current official eligibility rules.", "Owner": "User / CA"},
        {"Step": 5, "Action": "Prepare and validate the return in the current official utility or portal.", "Owner": "User / CA"},
        {"Step": 6, "Action": "Submit, e-verify and preserve the acknowledgement and records.", "Owner": "User"},
    ]

    workbook_bytes = _build_workbook(
        metadata=metadata,
        account_rows=account_rows,
        transaction_rows=transaction_rows,
        income_rows=income_rows,
        expense_rows=expense_rows,
        transfer_rows=transfer_rows,
        exception_rows=exceptions,
        period_rows=period_rows,
        evidence_rows=evidence_rows,
        filing_rows=filing_rows,
    )
    csv_bytes = _csv_bytes(transaction_rows)
    reconciliation = {
        "schema_version": TAX_PACK_SCHEMA_VERSION,
        "financial_year": fy_label,
        "assessment_year": ay_label,
        "transaction_count": len(transaction_rows),
        "income_candidate_total_inr": float(total_income),
        "expense_review_total_inr": float(total_expense),
        "transfer_or_reversal_count": len(transfer_rows),
        "data_quality_exception_count": len(exceptions),
        "period_warning_count": period_warning_count,
        "duplicate_risk_count": sum(
            1 for row in transaction_rows if row["Duplicate risk"]
        ),
        "unclassified_count": sum(
            1 for transaction in transactions if not transaction.category
        ),
        "readiness": (
            "review_required"
            if exceptions or period_warning_count
            else "automated_checks_clear_external_reconciliation_required"
        ),
        "warning": (
            "GODFIN cannot determine the correct return or filing readiness from "
            "transaction data alone. Send this encrypted archive and its passphrase "
            "through separate channels."
        ),
    }
    reconciliation_bytes = json.dumps(
        reconciliation, indent=2, sort_keys=True
    ).encode("utf-8")
    guide_bytes = _filing_guide_pdf(
        fy_label=fy_label,
        ay_label=ay_label,
        generated_at=generated.isoformat(),
        exception_count=len(exceptions),
        period_warning_count=period_warning_count,
    )

    short_label = f"fy{start_year}-{str(start_year + 1)[-2:]}"
    files = {
        f"godfin_ca_{short_label}.xlsx": workbook_bytes,
        f"godfin_ca_{short_label}_transactions.csv": csv_bytes,
        "reconciliation_summary.json": reconciliation_bytes,
        f"filing_guide_{ay_label.lower().replace(' ', '').replace('-', '_')}_v{TAX_GUIDE_VERSION}.pdf": guide_bytes,
    }
    manifest = {
        "schema_version": TAX_PACK_SCHEMA_VERSION,
        "pack_kind": "godfin_indian_financial_year_tax_pack",
        "financial_year": fy_label,
        "assessment_year": ay_label,
        "period_start": start.isoformat(),
        "period_end_exclusive": end.isoformat(),
        "generated_at_utc": generated.isoformat(),
        "encryption": {
            "format": "WinZip AES",
            "algorithm": "AES-256",
            "passphrase_stored": False,
            "filenames_encrypted": False,
        },
        "privacy": {
            "raw_narration_included": False,
            "free_text_locally_redacted": True,
            "account_numbers_masked": True,
            "warning": (
                "Archive filenames are visible without the passphrase. Share the "
                "passphrase through a separate channel."
            ),
        },
        "summary": reconciliation,
        "files": {
            name: {
                "sha256": hashlib.sha256(content).hexdigest(),
                "bytes": len(content),
            }
            for name, content in files.items()
        },
        "official_references": [OFFICIAL_DOWNLOADS_URL, OFFICIAL_AIS_URL],
    }
    files["manifest.json"] = json.dumps(
        manifest, indent=2, sort_keys=True
    ).encode("utf-8")

    archive = io.BytesIO()
    with pyzipper.AESZipFile(
        archive,
        "w",
        compression=pyzipper.ZIP_DEFLATED,
        encryption=pyzipper.WZ_AES,
    ) as package:
        package.setpassword(passphrase.encode("utf-8"))
        package.setencryption(pyzipper.WZ_AES, nbits=256)
        for name, content in files.items():
            package.writestr(name, content)
    return archive.getvalue()
